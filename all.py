"""
Find correlations between CPC vote change and ALL demographic categories.
Ranks demographics by correlation strength to discover the strongest predictors.
Creates professional Excel report with all 3 demographic labels.
"""

import psycopg2
import pandas as pd
from scipy import stats
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

CONNECTION_STRING = "postgresql://neondb_owner:npg_nDCY0KAWtN3z@ep-patient-night-a5t2b28c-pooler.us-east-2.aws.neon.tech/neondb?sslmode=require&channel_binding=require"


def load_cpc_votes(excel_file="cpc_votes_by_election.xlsx"):
    """Load CPC vote data from Excel file."""
    df = pd.read_excel(excel_file)
    
    df = df.rename(columns={
        "Constituency ID": "constituency_id",
        "Riding Name": "riding_name",
        "Province": "province",
        "CPC % Change (2021→2025)": "cpc_change_21_25",
        "CPC % Change (2019→2025)": "cpc_change_19_25",
        "CPC % 2021": "cpc_pct_2021",
        "CPC % 2025": "cpc_pct_2025"
    })
    
    df = df[df["cpc_change_21_25"].notna()]
    print(f"✅ Loaded {len(df)} constituencies from CPC votes file")
    return df


def fetch_all_demographics():
    """
    Fetch ALL demographic categories and their rates for each constituency.
    Returns: (DataFrame with rates, dict of category info)
    """
    conn = psycopg2.connect(CONNECTION_STRING)
    cur = conn.cursor()
    
    # Get all category info - NO TRUNCATION
    cat_query = """
        SELECT 
            id,
            characteristics_id,
            category,
            subcategory,
            subsubcategory,
            description
        FROM demographic_categories
        WHERE is_province = false OR is_province IS NULL
        ORDER BY characteristics_id
    """
    cur.execute(cat_query)
    categories = cur.fetchall()
    
    # Build category lookup - store ALL labels
    category_info = {}
    for row in categories:
        cat_id, char_id, cat, subcat, subsubcat, desc = row
        
        category_info[char_id] = {
            "category_id": cat_id,
            "category": cat or "",           # Level 1 (broadest)
            "subcategory": subcat or "",     # Level 2
            "subsubcategory": subsubcat or "", # Level 3 (most specific)
            "description": desc or ""
        }
    
    print(f"📊 Found {len(category_info)} demographic categories")
    
    # Fetch all demographic data
    data_query = """
        SELECT 
            dg.constituency_id,
            dc.characteristics_id,
            dg.values
        FROM demographics_geographic dg
        JOIN demographic_categories dc ON dg.category_id = dc.id
        WHERE dg.is_constituency = true
          AND dg.constituency_id IS NOT NULL
          AND (dc.is_province = false OR dc.is_province IS NULL)
        ORDER BY dg.constituency_id
    """
    
    cur.execute(data_query)
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
    
    print(f"📥 Processing {len(rows)} demographic records...")
    
    # Organize: constituency_id -> {char_id: rateTotal}
    demo_data = {}
    for constituency_id, char_id, values in rows:
        if constituency_id not in demo_data:
            demo_data[constituency_id] = {"constituency_id": constituency_id}
        
        rate = values.get("rateTotal", None) if values else None
        if rate is not None:
            demo_data[constituency_id][f"demo_{char_id}"] = rate
    
    df = pd.DataFrame(list(demo_data.values()))
    print(f"✅ Loaded demographics for {len(df)} constituencies")
    
    return df, category_info


def calculate_all_correlations(merged_df, category_info):
    """Calculate correlations for ALL demographic categories."""
    results = []
    
    demo_cols = [col for col in merged_df.columns if col.startswith("demo_")]
    print(f"\n🔍 Analyzing {len(demo_cols)} demographic variables...")
    
    for col in demo_cols:
        char_id = col.replace("demo_", "")
        
        valid_data = merged_df[["cpc_change_21_25", col]].dropna()
        
        if len(valid_data) < 30:
            continue
        
        if valid_data[col].std() == 0:
            continue
        
        r, p_value = stats.pearsonr(valid_data[col], valid_data["cpc_change_21_25"])
        
        info = category_info.get(char_id, {})
        
        # Determine correlation strength
        abs_r = abs(r)
        if abs_r >= 0.5:
            strength = "Strong"
        elif abs_r >= 0.3:
            strength = "Moderate"
        elif abs_r >= 0.2:
            strength = "Weak"
        else:
            strength = "Very Weak"
        
        results.append({
            "Characteristics ID": char_id,
            "Category (Level 1)": info.get("category", ""),
            "Subcategory (Level 2)": info.get("subcategory", ""),
            "Subsubcategory (Level 3)": info.get("subsubcategory", ""),
            "Pearson r": round(r, 4),
            "p-value": p_value,
            "Significant (p<0.05)": "Yes" if p_value < 0.05 else "No",
            "Correlation Strength": strength,
            "R-squared": round(r ** 2, 4),
            "Sample Size (n)": len(valid_data),
            "Direction": "Positive" if r > 0 else "Negative"
        })
    
    df = pd.DataFrame(results)
    df = df.sort_values("Pearson r", key=abs, ascending=False)
    
    return df


def create_excel_report(results_df, output_file="correlation_report.xlsx"):
    """Create professional Excel report with multiple sheets."""
    
    wb = Workbook()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_red = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Conditional fills for correlation values
    strong_pos_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    weak_pos_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    strong_neg_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    weak_neg_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    def write_sheet(ws, df, header_fill):
        """Write DataFrame to worksheet with formatting."""
        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Color code the Pearson r column (column 5)
                if col_idx == 5 and isinstance(value, (int, float)):
                    if value >= 0.3:
                        cell.fill = strong_pos_fill
                    elif value > 0:
                        cell.fill = weak_pos_fill
                    elif value <= -0.3:
                        cell.fill = strong_neg_fill
                    elif value < 0:
                        cell.fill = weak_neg_fill
                
                # Format p-value in scientific notation
                if col_idx == 6 and isinstance(value, float):
                    cell.number_format = '0.00E+00'
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Auto-adjust column widths
        column_widths = {
            "Characteristics ID": 18,
            "Category (Level 1)": 45,
            "Subcategory (Level 2)": 40,
            "Subsubcategory (Level 3)": 40,
            "Pearson r": 12,
            "p-value": 14,
            "Significant (p<0.05)": 18,
            "Correlation Strength": 20,
            "R-squared": 12,
            "Sample Size (n)": 15,
            "Direction": 12
        }
        
        for col, header in enumerate(headers, 1):
            width = column_widths.get(header, 15)
            ws.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 1: All Significant Correlations (sorted by strength)
    ws1 = wb.active
    ws1.title = "All Significant"
    sig_df = results_df[results_df["Significant (p<0.05)"] == "Yes"].copy()
    sig_df = sig_df.sort_values("Pearson r", key=abs, ascending=False)
    write_sheet(ws1, sig_df, header_fill_blue)
    
    # Sheet 2: Positive Correlations Only
    ws2 = wb.create_sheet("Positive Correlations")
    pos_df = results_df[(results_df["Significant (p<0.05)"] == "Yes") & (results_df["Direction"] == "Positive")].copy()
    pos_df = pos_df.sort_values("Pearson r", ascending=False)
    write_sheet(ws2, pos_df, header_fill_green)
    
    # Sheet 3: Negative Correlations Only
    ws3 = wb.create_sheet("Negative Correlations")
    neg_df = results_df[(results_df["Significant (p<0.05)"] == "Yes") & (results_df["Direction"] == "Negative")].copy()
    neg_df = neg_df.sort_values("Pearson r", ascending=True)
    write_sheet(ws3, neg_df, header_fill_red)
    
    # Sheet 4: All Results (including non-significant)
    ws4 = wb.create_sheet("All Results")
    write_sheet(ws4, results_df, header_fill_blue)
    
    # Sheet 5: Summary Statistics
    ws5 = wb.create_sheet("Summary")
    
    summary_data = [
        ["CORRELATION ANALYSIS SUMMARY", ""],
        ["", ""],
        ["Total demographics analyzed", len(results_df)],
        ["Statistically significant (p<0.05)", len(sig_df)],
        ["Positive correlations (significant)", len(pos_df)],
        ["Negative correlations (significant)", len(neg_df)],
        ["", ""],
        ["STRONGEST POSITIVE CORRELATIONS", ""],
    ]
    
    # Add top 5 positive
    for idx, row in pos_df.head(5).iterrows():
        summary_data.append([
            f"  {row['Subsubcategory (Level 3)'] or row['Subcategory (Level 2)'] or row['Category (Level 1)']}",
            f"r = {row['Pearson r']:.4f}"
        ])
    
    summary_data.append(["", ""])
    summary_data.append(["STRONGEST NEGATIVE CORRELATIONS", ""])
    
    # Add top 5 negative
    for idx, row in neg_df.head(5).iterrows():
        summary_data.append([
            f"  {row['Subsubcategory (Level 3)'] or row['Subcategory (Level 2)'] or row['Category (Level 1)']}",
            f"r = {row['Pearson r']:.4f}"
        ])
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 8, 15]:  # Headers
                cell.font = Font(bold=True, size=12)
    
    ws5.column_dimensions['A'].width = 60
    ws5.column_dimensions['B'].width = 20
    
    # Save
    wb.save(output_file)
    print(f"\n✅ Saved professional Excel report to {output_file}")
    print(f"   📊 Sheet 1: All Significant ({len(sig_df)} rows)")
    print(f"   📈 Sheet 2: Positive Correlations ({len(pos_df)} rows)")
    print(f"   📉 Sheet 3: Negative Correlations ({len(neg_df)} rows)")
    print(f"   📋 Sheet 4: All Results ({len(results_df)} rows)")
    print(f"   📝 Sheet 5: Summary Statistics")


def print_top_correlations(results_df, top_n=20):
    """Print the top correlations to console."""
    
    sig_df = results_df[results_df["Significant (p<0.05)"] == "Yes"].copy()
    
    print("\n" + "=" * 100)
    print(f"TOP {top_n} STRONGEST CORRELATIONS (Statistically Significant)")
    print("=" * 100)
    
    for idx, (_, row) in enumerate(sig_df.head(top_n).iterrows(), 1):
        direction = "+" if row["Pearson r"] > 0 else ""
        label = row["Subsubcategory (Level 3)"] or row["Subcategory (Level 2)"] or row["Category (Level 1)"]
        print(f"{idx:>3}. r = {direction}{row['Pearson r']:.4f}  |  {label[:70]}")
    
    # Positive only
    print("\n" + "=" * 100)
    print("TOP 10 POSITIVE (Higher demographic rate = More CPC vote increase)")
    print("=" * 100)
    
    pos_df = sig_df[sig_df["Direction"] == "Positive"].head(10)
    for _, row in pos_df.iterrows():
        label = row["Subsubcategory (Level 3)"] or row["Subcategory (Level 2)"] or row["Category (Level 1)"]
        print(f"  r = +{row['Pearson r']:.4f}  |  {label[:70]}")
    
    # Negative only
    print("\n" + "=" * 100)
    print("TOP 10 NEGATIVE (Higher demographic rate = Less CPC vote increase)")
    print("=" * 100)
    
    neg_df = sig_df[sig_df["Direction"] == "Negative"].sort_values("Pearson r").head(10)
    for _, row in neg_df.iterrows():
        label = row["Subsubcategory (Level 3)"] or row["Subcategory (Level 2)"] or row["Category (Level 1)"]
        print(f"  r = {row['Pearson r']:.4f}  |  {label[:70]}")
    
    return sig_df


def create_visualizations(results_df):
    """Create bar chart of top correlations."""
    
    sig_df = results_df[results_df["Significant (p<0.05)"] == "Yes"].copy()
    
    # Get labels (use most specific available)
    def get_label(row):
        return row["Subsubcategory (Level 3)"] or row["Subcategory (Level 2)"] or row["Category (Level 1)"]
    
    sig_df["Label"] = sig_df.apply(get_label, axis=1)
    
    top_pos = sig_df[sig_df["Direction"] == "Positive"].head(12)
    top_neg = sig_df[sig_df["Direction"] == "Negative"].sort_values("Pearson r").head(12)
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 10))
    
    # Positive correlations
    if len(top_pos) > 0:
        y_pos = range(len(top_pos))
        ax1.barh(y_pos, top_pos["Pearson r"], color="#2E7D32", alpha=0.85)
        ax1.set_yticks(y_pos)
        ax1.set_yticklabels([lab[:45] for lab in top_pos["Label"]], fontsize=9)
        ax1.set_xlabel("Pearson r", fontsize=11)
        ax1.set_title("Top Positive Correlations\n(Higher % → More CPC Vote Increase)", 
                      fontsize=12, fontweight="bold", color="#2E7D32")
        ax1.invert_yaxis()
        ax1.grid(axis="x", alpha=0.3)
        ax1.set_xlim(0, max(top_pos["Pearson r"]) * 1.1)
    
    # Negative correlations
    if len(top_neg) > 0:
        y_neg = range(len(top_neg))
        ax2.barh(y_neg, top_neg["Pearson r"], color="#C62828", alpha=0.85)
        ax2.set_yticks(y_neg)
        ax2.set_yticklabels([lab[:45] for lab in top_neg["Label"]], fontsize=9)
        ax2.set_xlabel("Pearson r", fontsize=11)
        ax2.set_title("Top Negative Correlations\n(Higher % → Less CPC Vote Increase)", 
                      fontsize=12, fontweight="bold", color="#C62828")
        ax2.invert_yaxis()
        ax2.grid(axis="x", alpha=0.3)
        ax2.set_xlim(min(top_neg["Pearson r"]) * 1.1, 0)
    
    plt.suptitle("Demographics Most Correlated with CPC Vote Change (2021→2025)", 
                 fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig("top_correlations.png", dpi=150, bbox_inches="tight")
    print(f"✅ Saved visualization to top_correlations.png")
    plt.show()


def main():
    print("=" * 100)
    print("COMPREHENSIVE DEMOGRAPHIC CORRELATION ANALYSIS")
    print("Finding which demographics correlate most with CPC vote change (2021→2025)")
    print("=" * 100)
    
    # Step 1: Load CPC votes
    print("\n📥 Loading CPC vote data...")
    cpc_df = load_cpc_votes()
    
    # Step 2: Fetch ALL demographics
    print("\n📥 Fetching ALL demographic data from database...")
    demo_df, category_info = fetch_all_demographics()
    
    # Step 3: Merge
    print("\n🔗 Merging datasets...")
    merged_df = pd.merge(cpc_df, demo_df, on="constituency_id", how="inner")
    print(f"✅ Merged: {len(merged_df)} constituencies")
    
    # Step 4: Calculate ALL correlations
    print("\n📈 Calculating correlations for all demographics...")
    results_df = calculate_all_correlations(merged_df, category_info)
    
    # Step 5: Print top correlations
    sig_df = print_top_correlations(results_df, top_n=20)
    
    # Step 6: Summary stats
    total_tested = len(results_df)
    significant = len(results_df[results_df["Significant (p<0.05)"] == "Yes"])
    pos_count = len(results_df[(results_df["Significant (p<0.05)"] == "Yes") & (results_df["Direction"] == "Positive")])
    neg_count = len(results_df[(results_df["Significant (p<0.05)"] == "Yes") & (results_df["Direction"] == "Negative")])
    
    print(f"\n" + "=" * 100)
    print("SUMMARY STATISTICS")
    print("=" * 100)
    print(f"   Total demographics tested: {total_tested}")
    print(f"   Statistically significant: {significant} ({100*significant/total_tested:.1f}%)")
    print(f"   Positive correlations: {pos_count}")
    print(f"   Negative correlations: {neg_count}")
    
    if len(results_df[results_df["Direction"] == "Positive"]) > 0:
        print(f"   Strongest positive r: {results_df[results_df['Direction'] == 'Positive']['Pearson r'].max():.4f}")
    if len(results_df[results_df["Direction"] == "Negative"]) > 0:
        print(f"   Strongest negative r: {results_df[results_df['Direction'] == 'Negative']['Pearson r'].min():.4f}")
    
    # Step 7: Create Excel report
    print("\n📊 Creating Excel report...")
    create_excel_report(results_df)
    
    # Step 8: Create visualizations
    print("\n📊 Creating visualizations...")
    create_visualizations(results_df)
    
    return merged_df, results_df


if __name__ == "__main__":
    merged_df, results = main()
