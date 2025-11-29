import json
from pathlib import Path
import psycopg2

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Run: pip install openpyxl")

CONNECTION_STRING = "postgresql://neondb_owner:npg_nDCY0KAWtN3z@ep-patient-night-a5t2b28c-pooler.us-east-2.aws.neon.tech/neondb?sslmode=require&channel_binding=require"


def extract_constituency_demographics(characteristics_ids, output_file="filtered_demographics.xlsx"):
    """
    Extract constituency-level demographics by characteristics_id from Neon database.
    
    Output format: Excel file with 4 sheets:
    - Sheet 1: Category 2255 (Sales and service occupations)
    - Sheet 2: Category 2256 (Trades, transport and equipment operators)
    - Sheet 3: Category 2257 (Natural resources, agriculture)
    - Sheet 4: SUM of all 3 categories
    
    Args:
        characteristics_ids: List of characteristics IDs to filter (as strings)
        output_file: Output Excel filename (.xlsx)
    """
    if not EXCEL_AVAILABLE:
        print("❌ Cannot create Excel file - openpyxl not installed")
        print("   Run: pip install openpyxl")
        return None
        
    conn = psycopg2.connect(CONNECTION_STRING)
    cur = conn.cursor()

    # Join demographics_geographic with demographic_categories
    query = """
        SELECT 
            dg.id,
            dg.constituency_id,
            dg.census_id,
            dg.category_id,
            dg.geographics_id,
            dg.is_constituency,
            dc.characteristics_id,
            dc.category,
            dc.subcategory,
            dc.subsubcategory,
            dc.is_province,
            dc.description,
            dg.values,
            dg.metadata
        FROM demographics_geographic dg
        JOIN demographic_categories dc ON dg.category_id = dc.id
        WHERE dc.characteristics_id = ANY(%s)
          AND (dc.is_province = false OR dc.is_province IS NULL)
          AND dg.is_constituency = true
        ORDER BY dc.characteristics_id, dg.constituency_id
    """

    cur.execute(query, (characteristics_ids,))
    rows = cur.fetchall()

    # Group data by constituency_id
    constituency_data = {}
    category_labels = {}  # Store labels for each characteristics_id
    
    for row in rows:
        constituency_id = row[1]
        characteristics_id = row[6]
        
        if constituency_id not in constituency_data:
            metadata = row[13] if row[13] else {}
            constituency_data[constituency_id] = {
                'base_info': {
                    'constituency_id': constituency_id,
                    'geoName': metadata.get('geoName', ''),
                    'sgcCode': metadata.get('sgcCode', ''),
                    'census_id': row[2],
                    'geographics_id': row[4],
                },
                'characteristics': {}
            }
        
        # Store values for each characteristic
        values = row[12] if row[12] else {}
        constituency_data[constituency_id]['characteristics'][characteristics_id] = {
            'subsubcategory': row[9],
            'total': values.get('total', 0) or 0,
            'men': values.get('men', 0) or 0,
            'women': values.get('women', 0) or 0,
            'rateTotal': values.get('rateTotal', 0) or 0,
            'rateMen': values.get('rateMen', 0) or 0,
            'rateWomen': values.get('rateWomen', 0) or 0
        }
        
        # Store the label for this characteristics_id
        if characteristics_id not in category_labels:
            category_labels[characteristics_id] = row[9]  # subsubcategory

    cur.close()
    conn.close()

    # Create Excel workbook with 4 sheets
    wb = Workbook()
    
    # Define headers
    headers = [
        "Constituency ID",
        "Riding Name",
        "SGC Code",
        "Census ID",
        "Geographic ID",
        "Total",
        "Men",
        "Women",
        "Rate Total (%)",
        "Rate Men (%)",
        "Rate Women (%)"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    sum_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    sum_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet names mapping
    sheet_names = {
        "2255": "Sales & Service",
        "2256": "Trades & Transport",
        "2257": "Natural Resources",
        "SUM": "SUM - All Categories"
    }
    
    # Create sheets for each category + SUM
    sheets_to_create = list(characteristics_ids) + ["SUM"]
    
    for idx, char_id in enumerate(sheets_to_create):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_names.get(char_id, f"Category {char_id}")
        else:
            ws = wb.create_sheet(title=sheet_names.get(char_id, f"Category {char_id}"))
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Write data rows
        row_num = 2
        for constituency_id in sorted(constituency_data.keys()):
            data = constituency_data[constituency_id]
            base = data['base_info']
            chars = data['characteristics']
            
            if char_id == "SUM":
                # Calculate sums
                sum_total = sum(chars[c]['total'] for c in chars)
                sum_men = sum(chars[c]['men'] for c in chars)
                sum_women = sum(chars[c]['women'] for c in chars)
                sum_rateTotal = sum(chars[c]['rateTotal'] for c in chars)
                sum_rateMen = sum(chars[c]['rateMen'] for c in chars)
                sum_rateWomen = sum(chars[c]['rateWomen'] for c in chars)
                
                row_data = [
                    base['constituency_id'],
                    base['geoName'],
                    base['sgcCode'],
                    base['census_id'],
                    base['geographics_id'],
                    sum_total,
                    sum_men,
                    sum_women,
                    round(sum_rateTotal, 1),
                    round(sum_rateMen, 1),
                    round(sum_rateWomen, 1)
                ]
                
                # Write with SUM styling
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.font = sum_font
                    cell.border = thin_border
            else:
                # Individual category data
                if char_id in chars:
                    char_data = chars[char_id]
                    row_data = [
                        base['constituency_id'],
                        base['geoName'],
                        base['sgcCode'],
                        base['census_id'],
                        base['geographics_id'],
                        char_data['total'],
                        char_data['men'],
                        char_data['women'],
                        char_data['rateTotal'],
                        char_data['rateMen'],
                        char_data['rateWomen']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
            
            row_num += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col-1]))
            for row in range(2, row_num):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Save workbook
    wb.save(output_file)
    
    unique_constituencies = len(constituency_data)
    
    print(f"✅ Exported {unique_constituencies} constituencies to {output_file}")
    print(f"   Sheets created:")
    for char_id in sheets_to_create:
        label = category_labels.get(char_id, "Sum of all categories")
        print(f"     - {sheet_names.get(char_id, char_id)}: {label}")
    
    return rows


if __name__ == "__main__":
    # Specify the characteristics IDs you want to extract
    # NOTE: characteristics_id is TEXT in database, so use strings
    target_ids = ["2255", "2256", "2257"]
    
    extract_constituency_demographics(target_ids)
