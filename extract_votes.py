import psycopg2

# For Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Run: pip install openpyxl")

CONNECTION_STRING = "postgresql://neondb_owner:npg_nDCY0KAWtN3z@ep-patient-night-a5t2b28c-pooler.us-east-2.aws.neon.tech/neondb?sslmode=require&channel_binding=require"


def extract_cpc_votes(output_file="cpc_votes_by_election.xlsx"):
    """
    Extract CPC (Conservative Party) votes for 2019, 2021, and 2025 federal elections.
    
    Output format: Excel file with columns for each election year.
    
    Database notes:
    - election_results.party_results is JSONB array with partyCode, votes, percentage
    - percentage is stored as DECIMAL (0-1), needs multiplication by 100
    - Elections are matched by election_id: 43 (2019), 44 (2021), 45 (2025)
    - rep_order = 2 means results mapped to 2023 redistribution boundaries
    - rep_order = 1 means original boundaries (used for 2025)
    """
    if not EXCEL_AVAILABLE:
        print("❌ Cannot create Excel file - openpyxl not installed")
        print("   Run: pip install openpyxl")
        return None
        
    conn = psycopg2.connect(CONNECTION_STRING)
    cur = conn.cursor()

    # Federal election IDs are known:
    # 43 = 2019, 44 = 2021, 45 = 2025
    # Verify these exist
    election_query = """
        SELECT e.id, EXTRACT(YEAR FROM e.start_date)::int as year, e.name
        FROM elections e
        WHERE e.id IN (43, 44, 45)
        ORDER BY e.start_date
    """
    
    cur.execute(election_query)
    elections = cur.fetchall()
    
    if not elections:
        print("❌ No federal elections found for IDs 43, 44, 45")
        cur.close()
        conn.close()
        return None
    
    print(f"📊 Found {len(elections)} federal elections:")
    election_map = {}  # year -> election_id
    for election_id, year, name in elections:
        print(f"   - {year}: {name} (ID: {election_id})")
        election_map[year] = election_id
    
    # Query election results with constituency info
    # For 2019/2021: use rep_order = 2 (2023 redistribution boundaries)
    # For 2025: use rep_order = 1 (current boundaries)
    # We'll take the highest rep_order available for each election to get consistent boundaries
    results_query = """
        SELECT 
            c.id as constituency_id,
            c.name as riding_name,
            c.code as riding_code,
            c.subnational as province,
            c.region,
            c.subregion,
            EXTRACT(YEAR FROM e.start_date)::int as election_year,
            er.total_votes,
            er.party_results,
            er.rep_order
        FROM election_results er
        JOIN constituencies c ON er.constituency_id = c.id
        JOIN elections e ON er.election_id = e.id
        WHERE e.id IN (43, 44, 45)
        ORDER BY c.id, e.start_date
    """
    
    cur.execute(results_query)
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
    
    if not rows:
        print("❌ No election results found")
        return None
    
    print(f"📊 Processing {len(rows)} election result records...")
    
    # Organize data by constituency
    # For each constituency+election, we may have multiple rep_orders
    # We want rep_order = 2 (2023 redistribution) for 2019/2021 to match current boundaries
    # For 2025, rep_order = 1 is fine
    constituency_data = {}
    
    for row in rows:
        constituency_id = row[0]
        riding_name = row[1]
        riding_code = row[2]
        province = row[3]
        region = row[4]
        subregion = row[5]
        election_year = row[6]
        total_votes = row[7]
        party_results = row[8]  # JSONB array
        rep_order = row[9]
        
        if constituency_id not in constituency_data:
            constituency_data[constituency_id] = {
                'riding_name': riding_name,
                'riding_code': riding_code,
                'province': province,
                'region': region,
                'subregion': subregion,
                'elections': {}
            }
        
        # For 2019/2021: prefer rep_order = 2 (2023 redistribution)
        # For 2025: use whatever is available (typically rep_order = 1)
        # Skip if we already have higher rep_order for this election
        existing = constituency_data[constituency_id]['elections'].get(election_year)
        if existing and existing.get('rep_order', 0) >= rep_order:
            continue  # Already have better or equal data
        
        # Extract CPC data from party_results JSONB
        cpc_votes = 0
        cpc_percentage = 0.0
        
        if party_results:
            for party in party_results:
                # CPC could be 'CPC' or 'Conservative' - check partyCode
                if party.get('partyCode') == 'CPC':
                    cpc_votes = party.get('votes', 0)
                    # IMPORTANT: percentage stored as decimal (0-1), multiply by 100
                    cpc_percentage = party.get('percentage', 0) * 100
                    break
        
        constituency_data[constituency_id]['elections'][election_year] = {
            'total_votes': total_votes,
            'cpc_votes': cpc_votes,
            'cpc_percentage': round(cpc_percentage, 2),
            'rep_order': rep_order
        }
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CPC Votes 2019-2021-2025"
    
    # Headers
    headers = [
        "Constituency ID",
        "Riding Name",
        "Riding Code",
        "Province",
        "Region",
        "Subregion",
        "CPC Votes 2019",
        "CPC % 2019",
        "Total Votes 2019",
        "CPC Votes 2021",
        "CPC % 2021",
        "Total Votes 2021",
        "CPC Votes 2025",
        "CPC % 2025",
        "Total Votes 2025",
        "CPC % Change (2021→2025)",
        "CPC % Change (2019→2025)"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Positive/negative change fills
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Write data rows
    row_num = 2
    for constituency_id in sorted(constituency_data.keys()):
        data = constituency_data[constituency_id]
        elections = data['elections']
        
        # Get data for each year (might be missing)
        e2019 = elections.get(2019, {})
        e2021 = elections.get(2021, {})
        e2025 = elections.get(2025, {})
        
        # Calculate changes
        pct_2019 = e2019.get('cpc_percentage', 0)
        pct_2021 = e2021.get('cpc_percentage', 0)
        pct_2025 = e2025.get('cpc_percentage', 0)
        
        change_21_25 = round(pct_2025 - pct_2021, 2) if pct_2021 and pct_2025 else None
        change_19_25 = round(pct_2025 - pct_2019, 2) if pct_2019 and pct_2025 else None
        
        row_data = [
            constituency_id,
            data['riding_name'],
            data['riding_code'],
            data['province'],
            data['region'],
            data['subregion'],
            e2019.get('cpc_votes', ''),
            pct_2019 if pct_2019 else '',
            e2019.get('total_votes', ''),
            e2021.get('cpc_votes', ''),
            pct_2021 if pct_2021 else '',
            e2021.get('total_votes', ''),
            e2025.get('cpc_votes', ''),
            pct_2025 if pct_2025 else '',
            e2025.get('total_votes', ''),
            change_21_25 if change_21_25 is not None else '',
            change_19_25 if change_19_25 is not None else ''
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Color-code the change columns
            if col in [16, 17] and value != '':  # Change columns
                if value > 0:
                    cell.fill = positive_fill
                elif value < 0:
                    cell.fill = negative_fill
        
        row_num += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col-1]))
        for row in range(2, min(row_num, 50)):  # Sample first 50 rows for width
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    
    # Summary statistics
    constituencies_with_all_years = sum(
        1 for c in constituency_data.values() 
        if 2019 in c['elections'] and 2021 in c['elections'] and 2025 in c['elections']
    )
    
    print(f"\n✅ Exported {len(constituency_data)} constituencies to {output_file}")
    print(f"   - Constituencies with all 3 elections: {constituencies_with_all_years}")
    print(f"   - Elections found: {sorted(election_map.keys())}")
    
    return constituency_data


def extract_cpc_votes_csv(output_file="cpc_votes_by_election.csv"):
    """
    CSV alternative if Excel isn't available.
    Same data, simpler format.
    """
    import csv
    
    conn = psycopg2.connect(CONNECTION_STRING)
    cur = conn.cursor()

    # Query election results with constituency info
    # Election IDs: 43 (2019), 44 (2021), 45 (2025)
    results_query = """
        SELECT 
            c.id as constituency_id,
            c.name as riding_name,
            c.code as riding_code,
            c.subnational as province,
            EXTRACT(YEAR FROM e.start_date)::int as election_year,
            er.total_votes,
            er.party_results,
            er.rep_order
        FROM election_results er
        JOIN constituencies c ON er.constituency_id = c.id
        JOIN elections e ON er.election_id = e.id
        WHERE e.id IN (43, 44, 45)
        ORDER BY c.id, e.start_date
    """
    
    cur.execute(results_query)
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
    
    # Organize by constituency
    # Prefer higher rep_order (2023 redistribution) for consistent boundaries
    constituency_data = {}
    
    for row in rows:
        constituency_id = row[0]
        riding_name = row[1]
        riding_code = row[2]
        province = row[3]
        election_year = row[4]
        total_votes = row[5]
        party_results = row[6]
        rep_order = row[7]
        
        if constituency_id not in constituency_data:
            constituency_data[constituency_id] = {
                'riding_name': riding_name,
                'riding_code': riding_code,
                'province': province,
                'elections': {}
            }
        
        # Skip if we already have higher rep_order for this election
        existing = constituency_data[constituency_id]['elections'].get(election_year)
        if existing and existing.get('rep_order', 0) >= rep_order:
            continue
        
        # Extract CPC data
        cpc_votes = 0
        cpc_percentage = 0.0
        
        if party_results:
            for party in party_results:
                if party.get('partyCode') == 'CPC':
                    cpc_votes = party.get('votes', 0)
                    cpc_percentage = party.get('percentage', 0) * 100
                    break
        
        constituency_data[constituency_id]['elections'][election_year] = {
            'total_votes': total_votes,
            'cpc_votes': cpc_votes,
            'cpc_percentage': round(cpc_percentage, 2),
            'rep_order': rep_order
        }
    
    # Write CSV
    headers = [
        "constituency_id", "riding_name", "riding_code", "province",
        "cpc_votes_2019", "cpc_pct_2019", "total_votes_2019",
        "cpc_votes_2021", "cpc_pct_2021", "total_votes_2021",
        "cpc_votes_2025", "cpc_pct_2025", "total_votes_2025",
        "pct_change_21_25", "pct_change_19_25"
    ]
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for constituency_id in sorted(constituency_data.keys()):
            data = constituency_data[constituency_id]
            elections = data['elections']
            
            e2019 = elections.get(2019, {})
            e2021 = elections.get(2021, {})
            e2025 = elections.get(2025, {})
            
            pct_2019 = e2019.get('cpc_percentage', 0)
            pct_2021 = e2021.get('cpc_percentage', 0)
            pct_2025 = e2025.get('cpc_percentage', 0)
            
            change_21_25 = round(pct_2025 - pct_2021, 2) if pct_2021 and pct_2025 else ''
            change_19_25 = round(pct_2025 - pct_2019, 2) if pct_2019 and pct_2025 else ''
            
            writer.writerow([
                constituency_id,
                data['riding_name'],
                data['riding_code'],
                data['province'],
                e2019.get('cpc_votes', ''),
                pct_2019 if pct_2019 else '',
                e2019.get('total_votes', ''),
                e2021.get('cpc_votes', ''),
                pct_2021 if pct_2021 else '',
                e2021.get('total_votes', ''),
                e2025.get('cpc_votes', ''),
                pct_2025 if pct_2025 else '',
                e2025.get('total_votes', ''),
                change_21_25,
                change_19_25
            ])
    
    print(f"✅ Exported {len(constituency_data)} constituencies to {output_file}")
    return constituency_data


if __name__ == "__main__":
    # Try Excel first, fall back to CSV
    if EXCEL_AVAILABLE:
        extract_cpc_votes()
    else:
        extract_cpc_votes_csv()

